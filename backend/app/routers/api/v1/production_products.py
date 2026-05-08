"""Production products router (產品管理)."""

import io
from typing import Optional, List
from decimal import Decimal, InvalidOperation

from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func

from app.core.database import get_db
from app.models.production import ProdProduct, ProdProductPackConfig
from app.models.user import User
from app.schemas.production import (
    ProdProductCreate,
    ProdProductUpdate,
    ProdProductResponse,
    FormingOption,
    ProdProductPackConfigUpsert,
    ProdProductPackConfigRead,
)
from app.schemas.common import PaginatedResponse
from app.dependencies.auth import get_current_active_user, require_role

router = APIRouter(prefix="/production/products", tags=["Production Products"])


def _to_response(product: ProdProduct, pack_config_count: int = 0) -> ProdProductResponse:
    return ProdProductResponse(
        id=product.id,
        code=product.code,
        name=product.name,
        ccp_limit_temp=product.ccp_limit_temp,
        pack_size_kg=product.pack_size_kg,
        loss_rate_warn_pct=product.loss_rate_warn_pct,
        product_type=product.product_type.value if hasattr(product.product_type, "value") else (product.product_type or "forming"),
        inv_item_id=product.inv_item_id,
        is_active=product.is_active,
        created_at=product.created_at,
        pack_config_count=pack_config_count,
    )


_PROD_SORT_MAP = {
    "code": ProdProduct.code,
    "name": ProdProduct.name,
    "product_type": ProdProduct.product_type,
}


@router.get("", response_model=PaginatedResponse[ProdProductResponse])
async def list_products(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    search: Optional[str] = None,
    show_inactive: bool = False,
    sort_by: str = Query("code", description="code | name | product_type"),
    sort_order: str = Query("asc", description="asc | desc"),
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    q = select(ProdProduct)
    if not show_inactive:
        q = q.where(ProdProduct.is_active == True)  # noqa: E712
    if search:
        q = q.where(
            ProdProduct.name.ilike(f"%{search}%")
            | ProdProduct.code.ilike(f"%{search}%")
        )

    total_result = await db.execute(
        select(func.count()).select_from(q.subquery())
    )
    total = total_result.scalar()

    sort_col = _PROD_SORT_MAP.get(sort_by, ProdProduct.code)
    order_expr = sort_col.desc() if sort_order == "desc" else sort_col.asc()
    items_result = await db.execute(
        q.order_by(order_expr).offset(skip).limit(limit)
    )
    items = items_result.scalars().all()

    # Single GROUP BY query for pack-config counts (avoids N+1)
    product_ids = [p.id for p in items]
    if product_ids:
        cnt_result = await db.execute(
            select(ProdProductPackConfig.product_id, func.count())
            .where(
                ProdProductPackConfig.product_id.in_(product_ids),
                ProdProductPackConfig.inv_item_id.is_not(None),
            )
            .group_by(ProdProductPackConfig.product_id)
        )
        pack_cfg_counts: dict[int, int] = {row[0]: row[1] for row in cnt_result}
    else:
        pack_cfg_counts = {}

    return PaginatedResponse(
        items=[_to_response(p, pack_cfg_counts.get(p.id, 0)) for p in items],
        total=total,
        skip=skip,
        limit=limit,
    )


@router.post(
    "", response_model=ProdProductResponse, status_code=status.HTTP_201_CREATED
)
async def create_product(
    data: ProdProductCreate,
    current_user: User = Depends(require_role("Admin", "Production")),
    db: AsyncSession = Depends(get_db),
):
    existing = await db.execute(
        select(ProdProduct).where(ProdProduct.code == data.code)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Product code '{data.code}' already exists",
        )
    product = ProdProduct(**data.model_dump())
    db.add(product)
    await db.flush()
    await db.commit()
    await db.refresh(product)
    return _to_response(product)


@router.get("/template")
async def download_template(
    current_user: User = Depends(get_current_active_user),
):
    """Download an Excel template for bulk product import."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "產品清單"

    headers = [
        ("產品代碼", "必填，不可重複"),
        ("產品名稱", "必填"),
        ("產品類型", "必填：forming（成型）或 hot_process（熱加工）"),
        ("CCP溫度限制°C", "選填，預設 75.00"),
        ("包裝規格kg", "選填"),
        ("損耗率警告%", "選填，0-100"),
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    note_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    example_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")

    # Row 1: headers
    for col, (header, _) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: notes
    for col, (_, note) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=note)
        cell.fill = note_fill
        cell.font = Font(italic=True, size=9, color="555555")

    # Row 3: example data
    examples = ["DC-PC-001", "豬肉水餃", "forming", "75.00", "1.0", "5.0"]
    for col, val in enumerate(examples, 1):
        cell = ws.cell(row=3, column=col, value=val)
        cell.fill = example_fill

    # Column widths
    widths = [18, 25, 20, 18, 15, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=prod_products_template.xlsx"},
    )


@router.post("/import")
async def import_products(
    file: UploadFile = File(...),
    current_user: User = Depends(require_role("Admin", "Production")),
    db: AsyncSession = Depends(get_db),
):
    """Bulk create production products from an uploaded Excel file."""
    import openpyxl

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="無法解析 Excel 檔案")

    rows = list(ws.iter_rows(min_row=3, values_only=True))  # skip header + notes

    created = 0
    skipped = 0
    errors: List[dict] = []

    VALID_TYPES = {"forming", "hot_process"}

    for i, row in enumerate(rows, start=3):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]
        code = row[0] if len(row) > 0 else ""
        name = row[1] if len(row) > 1 else ""
        product_type = row[2] if len(row) > 2 else "forming"
        ccp_raw = row[3] if len(row) > 3 else ""
        pack_raw = row[4] if len(row) > 4 else ""
        loss_raw = row[5] if len(row) > 5 else ""

        if not code:
            skipped += 1
            continue
        if not name:
            errors.append({"row": i, "code": code, "message": "產品名稱不可為空"})
            skipped += 1
            continue
        if product_type not in VALID_TYPES:
            product_type = "forming"

        # Check duplicate
        existing = await db.execute(select(ProdProduct).where(ProdProduct.code == code))
        if existing.scalar_one_or_none():
            errors.append({"row": i, "code": code, "message": f"代碼 '{code}' 已存在，略過"})
            skipped += 1
            continue

        def _dec(s: str):
            try:
                return Decimal(s) if s else None
            except InvalidOperation:
                return None

        product = ProdProduct(
            code=code,
            name=name,
            product_type=product_type,
            ccp_limit_temp=_dec(ccp_raw) or Decimal("75.00"),
            pack_size_kg=_dec(pack_raw),
            loss_rate_warn_pct=_dec(loss_raw),
        )
        db.add(product)
        created += 1

    await db.commit()
    return {"created": created, "skipped": skipped, "errors": errors}


@router.get("/forming-options", response_model=list[FormingOption])
async def get_forming_options(
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(ProdProduct)
        .where(ProdProduct.is_active == True)  # noqa: E712
        .order_by(ProdProduct.code)
    )
    products = result.scalars().all()
    return [
        FormingOption(
            id=p.id,
            code=p.code,
            name=p.name,
            product_type=p.product_type.value if hasattr(p.product_type, "value") else (p.product_type or "forming"),
            ccp_limit_temp=p.ccp_limit_temp,
            pack_size_kg=p.pack_size_kg,
            loss_rate_warn_pct=p.loss_rate_warn_pct,
        )
        for p in products
    ]


@router.get("/{product_id}", response_model=ProdProductResponse)
async def get_product(
    product_id: int,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(ProdProduct).where(ProdProduct.id == product_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Product not found"
        )
    return _to_response(product)


@router.patch("/{product_id}", response_model=ProdProductResponse)
async def update_product(
    product_id: int,
    data: ProdProductUpdate,
    current_user: User = Depends(require_role("Admin", "Production")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(ProdProduct).where(ProdProduct.id == product_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Product not found"
        )
    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(product, field, value)
    await db.flush()
    await db.commit()
    await db.refresh(product)
    return _to_response(product)


# ---------------------------------------------------------------------------
# Pack-config endpoints  GET / PUT  /production/products/{id}/pack-configs
# ---------------------------------------------------------------------------

@router.get("/{product_id}/pack-configs", response_model=List[ProdProductPackConfigRead])
async def get_pack_configs(
    product_id: int,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Return all pack-type → inv-item mappings for a product."""
    from sqlalchemy.orm import selectinload as _sil
    result = await db.execute(
        select(ProdProductPackConfig)
        .options(_sil(ProdProductPackConfig.inv_item))
        .where(ProdProductPackConfig.product_id == product_id)
        .order_by(ProdProductPackConfig.pack_type_code)
    )
    rows = result.scalars().all()
    return [
        ProdProductPackConfigRead(
            id=r.id,
            product_id=r.product_id,
            pack_type_code=r.pack_type_code,
            inv_item_id=r.inv_item_id,
            inv_item_code=r.inv_item.code if r.inv_item else None,
            inv_item_name=r.inv_item.name if r.inv_item else None,
        )
        for r in rows
    ]


@router.put("/{product_id}/pack-configs", response_model=List[ProdProductPackConfigRead])
async def save_pack_configs(
    product_id: int,
    configs: List[ProdProductPackConfigUpsert],
    current_user: User = Depends(require_role("Admin", "Warehouse", "Production")),
    db: AsyncSession = Depends(get_db),
):
    """Upsert the full list of pack-type configs for a product (merge, don't wipe)."""
    from sqlalchemy.orm import selectinload as _sil
    from sqlalchemy.dialects.postgresql import insert as pg_insert

    product = (await db.execute(select(ProdProduct).where(ProdProduct.id == product_id))).scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Product not found")

    for cfg in configs:
        existing = (await db.execute(
            select(ProdProductPackConfig)
            .where(ProdProductPackConfig.product_id == product_id)
            .where(ProdProductPackConfig.pack_type_code == cfg.pack_type_code)
        )).scalar_one_or_none()

        if existing:
            existing.inv_item_id = cfg.inv_item_id
        else:
            db.add(ProdProductPackConfig(
                product_id=product_id,
                pack_type_code=cfg.pack_type_code,
                inv_item_id=cfg.inv_item_id,
            ))

    await db.commit()

    result = await db.execute(
        select(ProdProductPackConfig)
        .options(_sil(ProdProductPackConfig.inv_item))
        .where(ProdProductPackConfig.product_id == product_id)
        .order_by(ProdProductPackConfig.pack_type_code)
    )
    rows = result.scalars().all()
    return [
        ProdProductPackConfigRead(
            id=r.id,
            product_id=r.product_id,
            pack_type_code=r.pack_type_code,
            inv_item_id=r.inv_item_id,
            inv_item_code=r.inv_item.code if r.inv_item else None,
            inv_item_name=r.inv_item.name if r.inv_item else None,
        )
        for r in rows
    ]
