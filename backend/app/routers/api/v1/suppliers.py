"""Suppliers router — CRUD + bulk-import for supplier management."""

import io
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func

from app.core.database import get_db
from app.models.supplier import Supplier
from app.models.user import User
from app.schemas.supplier import SupplierResponse, SupplierCreate, SupplierUpdate
from app.schemas.common import PaginatedResponse
from app.dependencies.auth import get_current_active_user, require_role

router = APIRouter(prefix="/suppliers", tags=["suppliers"])


@router.get("", response_model=PaginatedResponse[SupplierResponse])
async def list_suppliers(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    is_active: Optional[bool] = Query(None),
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """List suppliers with optional active filter."""
    query = select(Supplier)
    count_query = select(func.count(Supplier.id))
    if is_active is not None:
        query = query.where(Supplier.is_active == is_active)
        count_query = count_query.where(Supplier.is_active == is_active)

    total = (await db.execute(count_query)).scalar()
    result = await db.execute(query.order_by(Supplier.name).offset(skip).limit(limit))
    suppliers = result.scalars().all()

    return PaginatedResponse(
        items=[SupplierResponse.model_validate(s) for s in suppliers],
        total=total, skip=skip, limit=limit,
    )


@router.get("/template")
async def download_template(
    current_user: User = Depends(get_current_active_user),
):
    """Download an Excel template for bulk supplier import."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "供應商清單"

    headers = [
        ("供應商名稱", "必填，不可重複"),
        ("聯絡人", "選填"),
        ("電話", "選填"),
        ("電子郵件", "選填"),
        ("地址", "選填"),
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    note_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    example_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")

    for col, (header, _) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col, (_, note) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=note)
        cell.fill = note_fill
        cell.font = Font(italic=True, size=9, color="555555")

    examples = ["ABC Food Supply", "王小明", "0412345678", "abc@example.com", "123 Main St, Melbourne"]
    for col, val in enumerate(examples, 1):
        cell = ws.cell(row=3, column=col, value=val)
        cell.fill = example_fill

    widths = [30, 18, 18, 28, 40]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=suppliers_template.xlsx"},
    )


@router.post("/import")
async def import_suppliers(
    file: UploadFile = File(...),
    current_user: User = Depends(require_role("Admin", "QA")),
    db: AsyncSession = Depends(get_db),
):
    """Bulk create suppliers from an uploaded Excel file."""
    import openpyxl

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="無法解析 Excel 檔案")

    rows = list(ws.iter_rows(min_row=3, values_only=True))

    created = 0
    skipped = 0
    errors: List[dict] = []

    for i, row in enumerate(rows, start=3):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]
        name = row[0] if len(row) > 0 else ""
        contact_name = row[1] if len(row) > 1 else ""
        phone = row[2] if len(row) > 2 else ""
        email = row[3] if len(row) > 3 else ""
        address = row[4] if len(row) > 4 else ""

        if not name:
            skipped += 1
            continue

        existing = await db.execute(select(Supplier).where(Supplier.name == name))
        if existing.scalar_one_or_none():
            errors.append({"row": i, "code": name, "message": f"供應商 '{name}' 已存在，略過"})
            skipped += 1
            continue

        supplier = Supplier(
            name=name,
            contact_name=contact_name or None,
            phone=phone or None,
            email=email or None,
            address=address or None,
        )
        db.add(supplier)
        created += 1

    await db.commit()
    return {"created": created, "skipped": skipped, "errors": errors}


@router.get("/{supplier_id}", response_model=SupplierResponse)
async def get_supplier(
    supplier_id: int,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Get a single supplier."""
    result = await db.execute(select(Supplier).where(Supplier.id == supplier_id))
    supplier = result.scalar_one_or_none()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return SupplierResponse.model_validate(supplier)


@router.post("", response_model=SupplierResponse, status_code=status.HTTP_201_CREATED)
async def create_supplier(
    data: SupplierCreate,
    current_user: User = Depends(require_role("Admin", "QA")),
    db: AsyncSession = Depends(get_db),
):
    """Create a new supplier (Admin/QA only)."""
    supplier = Supplier(**data.model_dump())
    db.add(supplier)
    await db.commit()
    await db.refresh(supplier)
    return SupplierResponse.model_validate(supplier)


@router.patch("/{supplier_id}", response_model=SupplierResponse)
async def update_supplier(
    supplier_id: int,
    data: SupplierUpdate,
    current_user: User = Depends(require_role("Admin", "QA")),
    db: AsyncSession = Depends(get_db),
):
    """Update a supplier (Admin/QA only)."""
    result = await db.execute(select(Supplier).where(Supplier.id == supplier_id))
    supplier = result.scalar_one_or_none()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(supplier, field, value)

    await db.commit()
    await db.refresh(supplier)
    return SupplierResponse.model_validate(supplier)
