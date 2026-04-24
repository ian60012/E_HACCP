"""Inventory items router (品項管理)."""

import io
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.models.inventory import InvItem, InvLocation
from app.models.user import User
from app.schemas.inventory import (
    InvItemCreate, InvItemUpdate, InvItemResponse,
    InvAllowedLocationsUpdate, InvItemBulkUpdate,
)
from app.schemas.common import PaginatedResponse
from app.dependencies.auth import get_current_active_user, require_role

router = APIRouter(prefix="/inventory/items", tags=["inventory-items"])


def _to_response(item: InvItem) -> InvItemResponse:
    return InvItemResponse(
        id=item.id,
        code=item.code,
        name=item.name,
        category=item.category,
        base_unit=item.base_unit,
        description=item.description,
        supplier_id=item.supplier_id,
        supplier_name=item.supplier.name if item.supplier else None,
        is_active=item.is_active,
        created_at=item.created_at,
        allowed_location_ids=[loc.id for loc in item.allowed_locations],
    )


def _base_item_query():
    return select(InvItem).options(
        selectinload(InvItem.supplier),
        selectinload(InvItem.allowed_locations),
    )


async def _set_allowed_locations(db: AsyncSession, item: InvItem, location_ids: list[int]):
    """Replace item's allowed locations with the given list."""
    if location_ids:
        locs_result = await db.execute(
            select(InvLocation).where(InvLocation.id.in_(location_ids))
        )
        locs = locs_result.scalars().all()
        if len(locs) != len(location_ids):
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="One or more location IDs not found"
            )
        item.allowed_locations = locs
    else:
        item.allowed_locations = []


@router.get("", response_model=PaginatedResponse[InvItemResponse])
async def list_items(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    search: Optional[str] = None,
    is_active: Optional[bool] = True,
    category: Optional[str] = Query(None, description="Filter by category (e.g. '原料')"),
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    q = _base_item_query()
    if is_active is not None:
        q = q.where(InvItem.is_active == is_active)
    if search:
        q = q.where(
            InvItem.name.ilike(f"%{search}%") | InvItem.code.ilike(f"%{search}%")
        )
    if category:
        q = q.where(InvItem.category == category)

    total_result = await db.execute(select(func.count()).select_from(q.subquery()))
    total = total_result.scalar()

    items_result = await db.execute(q.order_by(InvItem.code).offset(skip).limit(limit))
    items = items_result.scalars().all()

    return PaginatedResponse(
        items=[_to_response(i) for i in items],
        total=total,
        skip=skip,
        limit=limit,
    )


@router.post("", response_model=InvItemResponse, status_code=status.HTTP_201_CREATED)
async def create_item(
    data: InvItemCreate,
    current_user: User = Depends(require_role("Admin", "Warehouse")),
    db: AsyncSession = Depends(get_db),
):
    existing = await db.execute(select(InvItem).where(InvItem.code == data.code))
    if existing.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Item code '{data.code}' already exists"
        )
    dump = data.model_dump(exclude={"allowed_location_ids"})
    item = InvItem(**dump)
    db.add(item)
    await db.flush()

    # Reload with relationships before assigning allowed_locations
    result = await db.execute(_base_item_query().where(InvItem.id == item.id))
    item = result.scalar_one()

    if data.allowed_location_ids is not None:
        await _set_allowed_locations(db, item, data.allowed_location_ids)
        await db.flush()

    await db.commit()
    result2 = await db.execute(_base_item_query().where(InvItem.id == item.id))
    return _to_response(result2.scalar_one())


@router.get("/template")
async def download_template(
    current_user: User = Depends(get_current_active_user),
):
    """Download an Excel template for bulk inventory item import."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "品項清單"

    headers = [
        ("品項代碼", "必填，不可重複"),
        ("品項名稱", "必填"),
        ("分類", "選填"),
        ("基本單位", "必填：PCS/KG/G/L/ML/包/箱/袋/罐/卷/打"),
        ("描述", "選填"),
        ("允許儲位", "選填，多個儲位以逗號分隔（儲位代碼）"),
        ("生產用量單位", "選填：KG/G/L/ML/PCS（Batch Sheet預設；空白=跟隨基本單位）"),
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    note_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    example_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")

    for col, (header, _) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col, (_, note) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=note)
        cell.fill = note_fill
        cell.font = Font(italic=True, size=9, color="555555")

    examples = ["RM-PORK-001", "豬前腿肉", "原料肉", "KG", "冷凍豬前腿肉", "FZ-A,FZ-B"]
    for col, val in enumerate(examples, 1):
        cell = ws.cell(row=3, column=col, value=val)
        cell.fill = example_fill

    widths = [18, 25, 15, 20, 25, 25]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inv_items_template.xlsx"},
    )


@router.post("/import")
async def import_items(
    file: UploadFile = File(...),
    current_user: User = Depends(require_role("Admin", "Warehouse")),
    db: AsyncSession = Depends(get_db),
):
    """Bulk create inventory items from an uploaded Excel file."""
    import openpyxl

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="無法解析 Excel 檔案")

    rows = list(ws.iter_rows(min_row=3, values_only=True))

    created = 0
    skipped = 0
    errors: List[dict] = []

    VALID_UNITS = {"PCS", "KG", "G", "L", "ML", "包", "箱", "袋", "罐", "卷", "打"}

    for i, row in enumerate(rows, start=3):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = [str(v).strip() if v is not None else "" for v in row]
        code = row[0] if len(row) > 0 else ""
        name = row[1] if len(row) > 1 else ""
        category = row[2] if len(row) > 2 else ""
        base_unit = row[3] if len(row) > 3 else "PCS"
        description = row[4] if len(row) > 4 else ""
        locations_raw = row[5] if len(row) > 5 else ""
        usage_unit = row[6] if len(row) > 6 else ""

        if not code:
            skipped += 1
            continue
        if not name:
            errors.append({"row": i, "code": code, "message": "品項名稱不可為空"})
            skipped += 1
            continue
        if not base_unit or base_unit not in VALID_UNITS:
            base_unit = "PCS"

        existing = await db.execute(select(InvItem).where(InvItem.code == code))
        if existing.scalar_one_or_none():
            errors.append({"row": i, "code": code, "message": f"代碼 '{code}' 已存在，略過"})
            skipped += 1
            continue

        item = InvItem(
            code=code,
            name=name,
            category=category or None,
            base_unit=base_unit,
            usage_unit=usage_unit or None,
            description=description or None,
        )
        db.add(item)
        await db.flush()

        # Resolve allowed locations by code
        if locations_raw:
            location_codes = [c.strip() for c in locations_raw.split(",") if c.strip()]
            if location_codes:
                locs_result = await db.execute(
                    select(InvLocation).where(InvLocation.code.in_(location_codes))
                )
                locs = locs_result.scalars().all()

                # Reload item with relationships to assign allowed_locations
                item_result = await db.execute(
                    select(InvItem)
                    .options(
                        selectinload(InvItem.supplier),
                        selectinload(InvItem.allowed_locations),
                    )
                    .where(InvItem.id == item.id)
                )
                item = item_result.scalar_one()
                item.allowed_locations = list(locs)
                await db.flush()

        created += 1

    await db.commit()
    return {"created": created, "skipped": skipped, "errors": errors}


@router.get("/{item_id}", response_model=InvItemResponse)
async def get_item(
    item_id: int,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(_base_item_query().where(InvItem.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    return _to_response(item)


@router.patch("/{item_id}", response_model=InvItemResponse)
async def update_item(
    item_id: int,
    data: InvItemUpdate,
    current_user: User = Depends(require_role("Admin", "Warehouse")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(_base_item_query().where(InvItem.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")

    dump = data.model_dump(exclude_unset=True, exclude={"allowed_location_ids"})
    for field, value in dump.items():
        setattr(item, field, value)

    if data.allowed_location_ids is not None:
        await _set_allowed_locations(db, item, data.allowed_location_ids)

    await db.flush()
    await db.commit()
    result2 = await db.execute(_base_item_query().where(InvItem.id == item_id))
    return _to_response(result2.scalar_one())


@router.patch("/bulk-update", status_code=status.HTTP_200_OK)
async def bulk_update_items(
    data: InvItemBulkUpdate,
    current_user: User = Depends(require_role("Admin", "Warehouse")),
    db: AsyncSession = Depends(get_db),
):
    """Bulk update category / base_unit / is_active for multiple items."""
    result = await db.execute(select(InvItem).where(InvItem.id.in_(data.ids)))
    items = result.scalars().all()

    for item in items:
        if data.category is not None:
            item.category = data.category or None
        if data.base_unit is not None:
            item.base_unit = data.base_unit
        if data.usage_unit is not None:
            item.usage_unit = data.usage_unit or None
        if data.is_active is not None:
            item.is_active = data.is_active

    await db.commit()
    return {"updated": len(items)}


@router.put("/{item_id}/allowed-locations", response_model=InvItemResponse)
async def set_allowed_locations(
    item_id: int,
    data: InvAllowedLocationsUpdate,
    current_user: User = Depends(require_role("Admin", "Warehouse")),
    db: AsyncSession = Depends(get_db),
):
    """Replace the allowed-location whitelist for an inventory item."""
    result = await db.execute(_base_item_query().where(InvItem.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")

    await _set_allowed_locations(db, item, data.location_ids)
    await db.flush()
    await db.commit()

    result2 = await db.execute(_base_item_query().where(InvItem.id == item_id))
    return _to_response(result2.scalar_one())
